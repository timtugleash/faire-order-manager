"""
Wholesale Order Manager — Streamlit App
=====================================
- Role-based login (admin / user)
- Pulls NEW and PROCESSING orders from Faire API
- WSP orders appear on Orders screen as PROCESSING
- PDF packing slip upload on WSP screen, download on Orders screen
- Downloads order data as Excel
- View Current Inventory from Google Sheets

SETUP:
  pip install streamlit requests openpyxl gspread google-auth google-api-python-client pandas

STREAMLIT SECRETS FORMAT:
  FAIRE_API_KEY = "..."
  ADMIN_PASSWORD = "..."
  USER_PASSWORD = "..."
  SHEET_ID = "..."
  DRIVE_FOLDER_ID = "..."  # Google Drive folder ID for packing slip PDFs

  [gcp_service_account]
  type = "service_account"
  project_id = "..."
  ...
"""

import io
import requests
import openpyxl
import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from google.auth.transport.requests import Request as GoogleAuthRequest
import base64
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
FAIRE_API_KEY    = st.secrets.get("FAIRE_API_KEY", "")
SHEET_ID         = st.secrets.get("SHEET_ID", "")
DRIVE_FOLDER_ID  = st.secrets.get("DRIVE_FOLDER_ID", "")

ALL_SKUS = [
    "T008-SBLK", "T008-MBLK", "T008-LBLK",
    "T008-SG",   "T008-MG",   "T008-LG",
    "T008-SC",   "T008-MC",   "T008-LC",
    "GRAB-H-SWT",  "GRAB-H-MWT",  "GRAB-H-LWT",  "GRAB-H-XLWT",
    "GRAB-H-SBLK", "GRAB-H-MBLK", "GRAB-H-LBLK", "GRAB-H-XLBLK",
    "GRAB-H-SG",   "GRAB-H-MG",   "GRAB-H-LG",   "GRAB-H-XLG",
    "GRAB-C-MWT",  "GRAB-C-LWT",  "GRAB-C-XLWT",
    "GRAB-C-MBLK", "GRAB-C-LBLK", "GRAB-C-XLBLK",
    "ROPE-OVL-BLK", "ROPE-OVL-BLU", "ROPE-OVL-GRN",
    "TUG-WM-MNY", "TUG-FL-02",
]

INCLUDE_STATES = {"NEW", "PROCESSING"}

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(page_title="Wholesale Order Manager", page_icon="📦", layout="wide")

# ─────────────────────────────────────────────
# ROLE-BASED LOGIN
# ─────────────────────────────────────────────
USERS = {
    "admin": {"password": st.secrets.get("ADMIN_PASSWORD", "shenzhen#1"), "role": "admin"},
    "jt":    {"password": st.secrets.get("USER_PASSWORD",  "tug2026"),    "role": "user"},
}

def login_screen():
    st.title("📦 Wholesale Order Manager")
    st.markdown("Please log in to continue.")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        user = USERS.get(username.lower())
        if user and password == user["password"]:
            st.session_state.authenticated = True
            st.session_state.role          = user["role"]
            st.session_state.username      = username.lower()
            st.rerun()
        else:
            st.error("Incorrect username or password.")
    st.stop()

if not st.session_state.get("authenticated"):
    login_screen()

role = st.session_state.get("role", "user")

# ─────────────────────────────────────────────
# GOOGLE SHEETS + DRIVE CONNECTION
# ─────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource
def get_credentials():
    return Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=SCOPES,
    )

@st.cache_resource
def get_gsheet_client():
    return gspread.authorize(get_credentials())

@st.cache_resource
def get_drive_service():
    return build("drive", "v3", credentials=get_credentials())


def get_sheet(tab_name: str):
    client = get_gsheet_client()
    sh     = client.open_by_key(SHEET_ID)
    return sh.worksheet(tab_name)


CHUNK_SIZE = 40000  # characters per cell, safely under 50000 limit

def store_pdf_in_sheet(pdf_bytes: bytes, filename: str) -> str:
    """Store PDF as chunked base64 across multiple rows in PDF_Store sheet."""
    b64    = base64.b64encode(pdf_bytes).decode("utf-8")
    chunks = [b64[i:i+CHUNK_SIZE] for i in range(0, len(b64), CHUNK_SIZE)]
    try:
        sh = get_gsheet_client().open_by_key(SHEET_ID)
        try:
            pdf_ws = sh.worksheet("PDF_Store")
        except Exception:
            pdf_ws = sh.add_worksheet(title="PDF_Store", rows=5000, cols=4)
            pdf_ws.append_row(["key", "filename", "chunk_index", "data"])

        for i, chunk in enumerate(chunks):
            pdf_ws.append_row([filename, filename, i, chunk])

        return filename
    except Exception as e:
        raise Exception(f"PDF store failed: {e}")


def retrieve_pdf_from_sheet(key: str) -> bytes:
    """Retrieve and reassemble a chunked PDF from PDF_Store sheet."""
    sh     = get_gsheet_client().open_by_key(SHEET_ID)
    pdf_ws = sh.worksheet("PDF_Store")
    rows   = pdf_ws.get_all_values()

    chunks = {}
    for row in rows[1:]:
        if len(row) >= 4 and row[0] == key:
            try:
                idx = int(row[2])
                chunks[idx] = row[3]
            except ValueError:
                pass

    if not chunks:
        raise Exception(f"PDF not found for key: {key}")

    b64 = "".join(chunks[i] for i in sorted(chunks.keys()))
    return base64.b64decode(b64)


def delete_pdf_from_sheet(key: str):
    """Delete all chunks for a PDF from PDF_Store sheet."""
    try:
        sh     = get_gsheet_client().open_by_key(SHEET_ID)
        pdf_ws = sh.worksheet("PDF_Store")
        rows   = pdf_ws.get_all_values()
        # Collect row indices to delete (in reverse to avoid index shifting)
        to_delete = [i + 2 for i, row in enumerate(rows[1:]) if len(row) >= 1 and row[0] == key]
        for row_idx in reversed(to_delete):
            pdf_ws.delete_rows(row_idx)
    except Exception:
        pass


def get_wsp_orders() -> list:
    """Fetch WSP orders from Google Sheets.
    Layout: rows = labels (Order Date, Order #, Customer, DriveFileID, SKU1...), columns = orders.
    Row 1 = Order Date, Row 2 = Order #, Row 3 = Customer, Row 4 = DriveFileID, Row 5+ = SKUs.
    Column A = labels, Column B onwards = one order per column.
    """
    try:
        ws   = get_sheet("WSP Orders")
        data = ws.get_all_values()  # data[row][col]

        if not data or len(data[0]) < 2:
            return []

        orders = []
        num_cols = max(len(row) for row in data)
        for col in range(1, num_cols):
            def cell(row_idx, c=col):
                try:
                    val = data[row_idx][c]
                    return val
                except IndexError:
                    return ""

            order_num = cell(1)  # Row 2 = Order #
            if not order_num:
                continue

            drive_file_id = cell(3)  # Row 4 = DriveFileID

            items = []
            for i, sku in enumerate(ALL_SKUS):
                row_idx = i + 4  # Row 5 onwards = SKUs (0-indexed: 4+)
                qty_str = cell(row_idx)
                try:
                    qty = int(qty_str)
                except (ValueError, TypeError):
                    qty = 0
                if qty > 0:
                    items.append({"sku": sku, "quantity": qty})

            orders.append({
                "order_number":  order_num,
                "raw_id":        f"wsp_{order_num}",
                "created_at":    cell(0),
                "state":         "NEW",
                "customer":      cell(2),
                "drive_file_id": drive_file_id,
                "items":         items,
                "source":        "WSP",
            })
        return orders
    except Exception as e:
        st.error(f"WSP fetch error: {e}")
        return []


# ─────────────────────────────────────────────
# FAIRE API FUNCTIONS
# ─────────────────────────────────────────────
def parse_order(data: dict) -> dict:
    items = []
    for item in data.get("items", []):
        sku = item.get("sku") or item.get("product_option", {}).get("sku", "UNKNOWN")
        items.append({"sku": sku, "quantity": item.get("quantity", 0)})

    created_raw = data.get("created_at", "")
    if isinstance(created_raw, (int, float)):
        created = datetime.utcfromtimestamp(created_raw / 1000).strftime("%Y-%m-%d")
    else:
        created = str(created_raw)[:10]

    return {
        "order_number": data.get("display_id", ""),
        "raw_id":       data.get("id", ""),
        "created_at":   created,
        "state":        data.get("state", ""),
        "customer":     data.get("address", {}).get("company_name", ""),
        "drive_file_id": "",
        "items":        items,
        "source":       "FAIRE",
    }


def fetch_faire_orders() -> list:
    headers = {"X-FAIRE-ACCESS-TOKEN": FAIRE_API_KEY}
    orders  = []
    cursor  = None

    while True:
        params = {"limit": 50, "sort_by": "CREATED_AT"}
        if cursor:
            params["cursor"] = cursor

        r = requests.get(
            "https://www.faire.com/external-api/v2/orders",
            headers=headers,
            params=params,
        )
        r.raise_for_status()
        data  = r.json()
        batch = data.get("orders", [])

        active  = [o for o in batch if o.get("state", "").upper() in INCLUDE_STATES]
        orders += [parse_order(o) for o in active]

        cursor = data.get("cursor")
        if not cursor or not batch:
            break

    return orders


def fetch_packing_slip(raw_id: str) -> bytes:
    headers = {"X-FAIRE-ACCESS-TOKEN": FAIRE_API_KEY}
    url     = f"https://www.faire.com/external-api/v2/orders/{raw_id}/packing-slip-pdf"
    r       = requests.get(url, headers=headers)
    r.raise_for_status()
    return r.content


# ─────────────────────────────────────────────
# GOOGLE SHEETS SYNC
# ─────────────────────────────────────────────

def get_existing_order_numbers_from_sheet() -> set:
    """Read order numbers already stored in the Faire Orders sheet tab."""
    try:
        ws   = get_sheet("Faire Orders")
        # Order # is in row 3 (index 2), columns B onwards
        row3 = ws.row_values(3)
        return set(v for v in row3[1:] if v)  # skip column A label
    except Exception:
        return set()


def sync_orders_to_sheet(orders: list):
    """Add new orders as columns to the Faire Orders sheet. Never overwrites existing."""
    try:
        ws               = get_sheet("Faire Orders")
        existing_nums    = get_existing_order_numbers_from_sheet()
        new_orders       = [o for o in orders if o["order_number"] not in existing_nums]

        if not new_orders:
            return 0

        data     = ws.get_all_values()
        max_cols = max((len(row) for row in data), default=0)

        from gspread.utils import rowcol_to_a1

        for order in new_orders:
            next_col   = max_cols + 1
            sku_lookup = {item["sku"]: item["quantity"] for item in order["items"]}

            col_values = [
                order["raw_id"],        # Row 1: raw_id (Faire internal ID)
                order["created_at"],    # Row 2: Order Date
                order["order_number"],  # Row 3: Order #
                order["customer"],      # Row 4: Customer
                "",                     # Row 5: blank separator
            ] + [
                sku_lookup.get(sku, "") for sku in ALL_SKUS  # Row 6+: SKUs
            ]

            cell_updates = []
            for row_idx, val in enumerate(col_values, start=1):
                if val != "":
                    cell_updates.append({
                        "range":  rowcol_to_a1(row_idx, next_col),
                        "values": [[val]],
                    })
            if cell_updates:
                ws.batch_update(cell_updates)
            max_cols += 1

        return len(new_orders)
    except Exception as e:
        st.warning(f"Could not sync to Google Sheets: {e}")
        return 0


def load_orders_from_sheet() -> list:
    """Load previously synced Faire orders from Google Sheets."""
    try:
        ws   = get_sheet("Faire Orders")
        data = ws.get_all_values()

        if not data or len(data[0]) < 2:
            return []

        orders    = []
        num_cols  = max(len(row) for row in data)

        for col in range(1, num_cols):
            def cell(row_idx, c=col):
                try:
                    return data[row_idx][c]
                except IndexError:
                    return ""

            order_num = cell(2)  # Row 3 = Order #
            if not order_num:
                continue

            items = []
            for i, sku in enumerate(ALL_SKUS):
                row_idx = i + 5  # Row 6+ = SKUs
                qty_str = cell(row_idx)
                try:
                    qty = int(qty_str)
                except (ValueError, TypeError):
                    qty = 0
                if qty > 0:
                    items.append({"sku": sku, "quantity": qty})

            orders.append({
                "order_number": order_num,
                "raw_id":       cell(0),   # Row 1 = raw_id
                "created_at":   cell(1),   # Row 2 = Order Date
                "state":        "NEW",
                "customer":     cell(3),   # Row 4 = Customer
                "drive_file_id": "",
                "items":        items,
                "source":       "FAIRE",
            })
        return orders
    except Exception as e:
        st.warning(f"Could not load orders from sheet: {e}")
        return []


# ─────────────────────────────────────────────
# FINALIZE ORDERS
# ─────────────────────────────────────────────

def get_existing_all_orders_numbers() -> set:
    """Get order numbers already in the All Orders tab.
    Row 1 contains 'Order# - Customer' so we extract the order number prefix."""
    try:
        ws   = get_sheet("All Orders")
        row1 = ws.row_values(1)
        nums = set()
        for val in row1[1:]:
            if val:
                # Extract order number before " - "
                nums.add(val.split(" - ")[0].strip())
        return nums
    except Exception:
        return set()


def finalize_orders_to_sheet(faire_orders: list, wsp_orders: list) -> tuple:
    """Copy all new + processing orders to All Orders tab as new columns."""
    from gspread.utils import rowcol_to_a1

    existing   = get_existing_all_orders_numbers()
    all_to_add = [o for o in faire_orders + wsp_orders if o["order_number"] not in existing]

    if not all_to_add:
        return 0, 0

    ws       = get_sheet("All Orders")
    data     = ws.get_all_values()
    max_cols = max((len(row) for row in data), default=0)

    # Expand sheet columns if needed
    needed_cols = max_cols + len(all_to_add) + 5
    if needed_cols > ws.col_count:
        ws.resize(cols=needed_cols)

    added_faire = 0
    added_wsp   = 0

    for order in all_to_add:
        next_col   = max_cols + 1
        sku_lookup = {item["sku"]: item["quantity"] for item in order["items"]}

        # Layout: Row1 = "Order# - Customer", Row2 = Date, Row3+ = SKUs
        order_label = f"{order['order_number']} - {order['customer']}"
        # Format date as mm/dd
        try:
            short_date = datetime.strptime(order["created_at"], "%Y-%m-%d").strftime("%m/%d")
        except Exception:
            short_date = order["created_at"]
        col_values  = [
            order_label,   # Row 1: Order # - Customer
            short_date,    # Row 2: Date as mm/dd
        ] + [
            sku_lookup.get(sku, "") for sku in ALL_SKUS  # Row 3+: SKUs
        ]

        cell_updates = []
        for row_idx, val in enumerate(col_values, start=1):
            if val != "":
                cell_updates.append({
                    "range":  rowcol_to_a1(row_idx, next_col),
                    "values": [[val]],
                })
        if cell_updates:
            ws.batch_update(cell_updates)

        max_cols += 1
        if order["source"] == "FAIRE":
            added_faire += 1
        else:
            added_wsp += 1

    return added_faire, added_wsp


# ─────────────────────────────────────────────
# SHEET EXPORT
# ─────────────────────────────────────────────

def sheet_to_excel(tab_name: str) -> bytes:
    """Export a Google Sheet tab as formatted Excel using Google export API."""
    client = get_gsheet_client()
    sh     = client.open_by_key(SHEET_ID)
    ws     = sh.worksheet(tab_name)
    gid    = ws.id

    export_url = (
        f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export"
        f"?format=xlsx&gid={gid}"
    )
    creds = get_credentials()
    creds.refresh(GoogleAuthRequest())
    headers  = {"Authorization": f"Bearer {creds.token}"}
    response = requests.get(export_url, headers=headers)
    response.raise_for_status()
    return response.content


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────
STORAGE_BOXES = {
    "T008-SBLK":    "RDL1",
    "T008-MBLK":    "RDL2",
    "T008-LBLK":    "RDL3",
    "T008-SG":      "RDL4",
    "T008-MG":      "RDL5",
    "T008-LG":      "RDL6",
    "T008-SC":      "RDL7",
    "T008-MC":      "RDL8",
    "T008-LC":      "RDL9",
    "GRAB-H-SWT":   "GH1",
    "GRAB-H-MWT":   "GH2",
    "GRAB-H-LWT":   "GH3",
    "GRAB-H-XLWT":  "GH4",
    "GRAB-H-SBLK":  "GH5",
    "GRAB-H-MBLK":  "GH6",
    "GRAB-H-LBLK":  "GH7",
    "GRAB-H-XLBLK": "GH8",
    "GRAB-H-SG":    "GH9",
    "GRAB-H-MG":    "GH10",
    "GRAB-H-LG":    "GH11",
    "GRAB-H-XLG":   "GH12",
    "GRAB-C-MWT":   "GC1",
    "GRAB-C-LWT":   "GC2",
    "GRAB-C-XLWT":  "GC3",
    "GRAB-C-MBLK":  "GC4",
    "GRAB-C-LBLK":  "GC5",
    "GRAB-C-XLBLK": "GC6",
    "ROPE-OVL-BLK": "RL1",
    "ROPE-OVL-BLU": "RL2",
    "ROPE-OVL-GRN": "RL3",
    "TUG-WM-MNY":   "",
    "TUG-FL-02":    "",
}


def build_excel(orders: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Data"

    lookup = {
        order["order_number"]: {item["sku"]: item["quantity"] for item in order["items"]}
        for order in orders
    }

    ROW_DATE, ROW_ORDER, ROW_CUSTOMER, ROW_BLANK, ROW_SKU_START = 1, 2, 3, 4, 5

    # Column A — "SKU" header + SKU labels
    for row, label in [(ROW_DATE, ""), (ROW_ORDER, ""), (ROW_CUSTOMER, ""), (ROW_BLANK, "SKU")]:
        cell = ws.cell(row=row, column=1, value=label)
        if label == "SKU":
            cell.font      = Font(bold=True, name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color="D9E1F2")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for i, sku in enumerate(ALL_SKUS):
        cell = ws.cell(row=ROW_SKU_START + i, column=1, value=sku)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column B — "Storage Box" header + storage box values
    storage_header = ws.cell(row=ROW_BLANK, column=2, value="Storage Box")
    storage_header.font      = Font(bold=True, name="Arial", size=10)
    storage_header.fill      = PatternFill("solid", start_color="D9E1F2")
    storage_header.alignment = Alignment(horizontal="left", vertical="center")

    for i, sku in enumerate(ALL_SKUS):
        cell = ws.cell(row=ROW_SKU_START + i, column=2, value=STORAGE_BOXES.get(sku, ""))
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Columns C onwards — one order per column
    for col_offset, order in enumerate(orders):
        col = col_offset + 3  # start at column C

        date_cell = ws.cell(row=ROW_DATE, column=col, value=order["created_at"])
        date_cell.font      = Font(name="Arial", size=10)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        ord_cell = ws.cell(row=ROW_ORDER, column=col, value=order["order_number"])
        ord_cell.font      = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        ord_cell.fill      = PatternFill("solid", start_color="2F5496")
        ord_cell.alignment = Alignment(horizontal="center", vertical="center")

        cust_cell = ws.cell(row=ROW_CUSTOMER, column=col, value=order["customer"])
        cust_cell.font      = Font(name="Arial", size=10)
        cust_cell.alignment = Alignment(horizontal="center", vertical="center")

        order_skus = lookup.get(order["order_number"], {})
        for i, sku in enumerate(ALL_SKUS):
            qty  = order_skus.get(sku, 0)
            cell = ws.cell(row=ROW_SKU_START + i, column=col, value=qty if qty else "")
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for col_offset in range(len(orders)):
        ws.column_dimensions[get_column_letter(col_offset + 3)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# SKU WEIGHTS
# ─────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def load_sku_weights() -> dict:
    """Load SKU weights from the SKU Weights tab in Google Sheets."""
    try:
        ws   = get_sheet("SKU Weights")
        rows = ws.get_all_values()
        weights = {}
        for row in rows[1:]:  # skip header
            if len(row) >= 2 and row[0] and row[1]:
                try:
                    weights[row[0].strip()] = float(row[1])
                except ValueError:
                    pass
        return weights
    except Exception as e:
        st.warning(f"Could not load SKU weights: {e}")
        return {}

def get_sku_weight(sku: str) -> float:
    return load_sku_weights().get(sku, 0.0)


# ─────────────────────────────────────────────
# SHIPPING INFO FUNCTIONS
# ─────────────────────────────────────────────

def save_carton(order_num: str, carton_num: int, length: float, width: float, height: float, sku_qtys: dict):
    """Save a carton's info to the Shipping Info sheet. One row per SKU."""
    ws  = get_sheet("Shipping Info")
    # Check if header exists
    existing = ws.get_all_values()
    if not existing or existing[0] != ["Order #", "Carton #", "Length", "Width", "Height", "SKU", "Qty", "Weight (lbs)"]:
        ws.insert_row(["Order #", "Carton #", "Length", "Width", "Height", "SKU", "Qty", "Weight (lbs)"], 1)

    for sku, qty in sku_qtys.items():
        if qty > 0:
            weight = round(get_sku_weight(sku) * qty, 3)
            ws.append_row([order_num, carton_num, length, width, height, sku, qty, weight])


def delete_carton(order_num: str, carton_num: int):
    """Delete all rows for a specific carton."""
    ws   = get_sheet("Shipping Info")
    rows = ws.get_all_values()
    # Collect row indices to delete in reverse
    to_delete = [
        i + 1 for i, row in enumerate(rows)
        if len(row) >= 2 and row[0] == order_num and str(row[1]) == str(carton_num)
    ]
    for row_idx in reversed(to_delete):
        ws.delete_rows(row_idx)


def get_shipping_info(order_num: str) -> dict:
    """Get all cartons for an order. Returns {carton_num: {dims, skus}}."""
    try:
        ws   = get_sheet("Shipping Info")
        rows = ws.get_all_values()
        cartons = {}
        for row in rows[1:]:
            if len(row) < 7 or row[0] != order_num:
                continue
            c_num = int(row[1])
            if c_num not in cartons:
                cartons[c_num] = {
                    "length": row[2], "width": row[3], "height": row[4],
                    "skus": {}
                }
            try:
                cartons[c_num]["skus"][row[5]] = int(row[6])
            except ValueError:
                pass
        return cartons
    except Exception:
        return {}


# ─────────────────────────────────────────────
# INVENTORY PUSH TO FAIRE
# ─────────────────────────────────────────────

def push_inventory_to_faire(inventories: list) -> dict:
    """Push inventory updates to Faire API.
    inventories: list of {"sku": str, "on_hand_quantity": int}
    """
    url     = "https://www.faire.com/external-api/v2/product-inventory/by-skus"
    headers = {
        "X-FAIRE-ACCESS-TOKEN": FAIRE_API_KEY,
        "Content-Type": "application/json",
    }
    payload  = {"inventories": inventories}
    response = requests.patch(url, headers=headers, json=payload)
    response.raise_for_status()
    return response.json()


# ─────────────────────────────────────────────
# HEADER + NAVIGATION
# ─────────────────────────────────────────────
col_title, col_logout = st.columns([6, 1])
with col_title:
    st.title("📦 Wholesale Order Manager")
    st.caption(f"Logged in as **{st.session_state.username}** ({role})")
with col_logout:
    st.write("")
    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()

pages = ["📋 Orders", "📊 Inventory", "📦 Shipping Info"]
if role == "admin":
    pages.append("🛒 WSP Orders")

page = st.sidebar.radio("Navigation", pages)
st.sidebar.divider()
st.sidebar.caption(f"Logged in as **{st.session_state.username}**")


# ─────────────────────────────────────────────
# PAGE: ORDERS
# ─────────────────────────────────────────────
if page == "📋 Orders":
    st.header("📋 New & Processing Orders")
    st.caption("Showing NEW and PROCESSING orders from Faire and WholesalePet.com.")

    if not FAIRE_API_KEY:
        st.error("No Faire API key found.")
        st.stop()

    # Load orders from Google Sheets on first visit
    if "faire_orders" not in st.session_state:
        with st.spinner("Loading orders from sheet..."):
            st.session_state["faire_orders"] = load_orders_from_sheet()

    # Cache controls
    btn_col1, btn_col2 = st.columns([2, 2])
    with btn_col2:
        if st.button("🗑️ Clear Local Cache"):
            if "faire_orders" in st.session_state:
                del st.session_state["faire_orders"]
            st.rerun()

    # Manual refresh button — pulls from Faire API and syncs new orders to sheet
    with btn_col1:
     if st.button("🔄 Refresh from Faire"):
        with st.spinner("Fetching latest orders from Faire..."):
            try:
                fresh_orders = fetch_faire_orders()
                added        = sync_orders_to_sheet(fresh_orders)
                # Reload from sheet to get full up-to-date list
                st.session_state["faire_orders"] = load_orders_from_sheet()
                if added:
                    st.success(f"✅ {added} new order(s) added to sheet!")
                else:
                    st.info("No new orders found — sheet is already up to date.")
            except Exception as e:
                st.error(f"Failed to fetch from Faire: {e}")

    faire_orders = st.session_state.get("faire_orders", [])
    wsp_orders   = get_wsp_orders()
    all_orders   = faire_orders + wsp_orders

    if not all_orders:
        st.info("No orders found. Click '🔄 Refresh from Faire' to pull latest orders.")
        st.stop()

    faire_count = len(faire_orders)
    wsp_count   = len(wsp_orders)
    st.success(f"**{len(all_orders)} order(s)** — {faire_count} from Faire, {wsp_count} from WholesalePet.com")

    excel_bytes = build_excel(all_orders)
    st.download_button(
        label     = "⬇️ Download Excel (New Orders)",
        data      = excel_bytes,
        file_name = f"new_orders_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Finalize Orders — double confirmation
    if role == "admin":
        st.divider()
        st.subheader("✅ Finalize Orders")
        st.caption("Moves all current orders to the All Orders tab in Google Sheets.")

        if "finalize_confirm1" not in st.session_state:
            st.session_state["finalize_confirm1"] = False

        if not st.session_state["finalize_confirm1"]:
            if st.button("✅ Finalize Orders", type="primary"):
                st.session_state["finalize_confirm1"] = True
                st.rerun()
        else:
            st.warning("⚠️ Are you sure? This will copy all current orders to the All Orders tab. This cannot be undone.")
            col_yes, col_no = st.columns([1, 1])
            with col_yes:
                if st.button("✅ Yes, Finalize", type="primary"):
                    with st.spinner("Finalizing orders..."):
                        try:
                            added_faire, added_wsp = finalize_orders_to_sheet(faire_orders, wsp_orders)
                            st.session_state["finalize_confirm1"] = False
                            # Clear orders from screen after finalizing
                            if "faire_orders" in st.session_state:
                                del st.session_state["faire_orders"]
                            st.success(f"✅ Done! {added_faire} Faire + {added_wsp} WSP order(s) added to All Orders tab.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to finalize: {e}")
                            st.session_state["finalize_confirm1"] = False
            with col_no:
                if st.button("❌ Cancel"):
                    st.session_state["finalize_confirm1"] = False
                    st.rerun()

    st.divider()

    for order in all_orders:
        customer_safe = (order["customer"] or "Unknown").replace("/", "-").replace("\\", "-")
        filename      = f"{order['order_number']}_{customer_safe}_PackingSlip.pdf"
        source_label  = "🛒 WSP" if order["source"] == "WSP" else "🏪 Faire"

        with st.container(border=True):
            st.markdown(f"**{order['order_number']}** — {order['customer'] or '—'}")
            st.caption(f"{order['created_at']}  |  {order['state']}  |  {source_label}")

            if order["source"] == "WSP":
                if order.get("drive_file_id"):
                    try:
                        pdf_bytes = retrieve_pdf_from_sheet(order["drive_file_id"])
                        st.download_button(
                            label     = "⬇️ Packing Slip PDF",
                            data      = pdf_bytes,
                            file_name = filename,
                            mime      = "application/pdf",
                            key       = f"pdf_{order['raw_id']}",
                        )
                    except Exception:
                        st.caption("PDF unavailable")
                else:
                    st.caption("No PDF uploaded")
            else:
                try:
                    pdf_bytes = fetch_packing_slip(order["raw_id"])
                    st.download_button(
                        label     = "⬇️ Packing Slip PDF",
                        data      = pdf_bytes,
                        file_name = filename,
                        mime      = "application/pdf",
                        key       = f"pdf_{order['raw_id']}",
                    )
                except Exception:
                    st.caption("PDF unavailable")


# ─────────────────────────────────────────────
# PAGE: INVENTORY
# ─────────────────────────────────────────────
elif page == "📊 Inventory":
    st.header("📊 Current Inventory")
    st.caption("Read-only view from Google Sheets.")

    if "inv_data" not in st.session_state:
        st.session_state["inv_data"] = None

    top_col1, top_col2 = st.columns([1, 1])
    with top_col1:
        if st.button("🔄 Refresh Inventory"):
            st.session_state["inv_data"] = None
    with top_col2:
        try:
            inv_excel = sheet_to_excel("Inventory")
            st.download_button(
                label     = "⬇️ Download Inventory",
                data      = inv_excel,
                file_name = "inventory.xlsx",
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Could not prepare Inventory download: {e}")



    if st.session_state["inv_data"] is None:
        with st.spinner("Loading inventory..."):
            try:
                client = get_gsheet_client()
                sh     = client.open_by_key(SHEET_ID)
                ws     = sh.worksheet("Inventory")
                rows   = ws.get_all_values()
                if rows and len(rows) > 1:
                    inv_rows = []
                    for row in rows[1:]:
                        if len(row) < 2 or not row[1]:
                            continue
                        inv_rows.append(row)
                    st.session_state["inv_data"] = inv_rows
                else:
                    st.session_state["inv_data"] = []
            except Exception as e:
                st.error(f"Could not load inventory: {e}")
                st.session_state["inv_data"] = []

    rows = st.session_state["inv_data"]
    try:
        if not rows:
            st.info("No inventory data found.")
        else:
            inv_data  = []
            for row in rows:
                if len(row) < 2 or not row[1]:
                    continue
                inv_data.append({
                    "Product":           row[0]  if len(row) > 0  else "",
                    "SKU":               row[1]  if len(row) > 1  else "",
                    "Storage Box":       row[2]  if len(row) > 2  else "",
                    "Pcs/Carton":        row[3]  if len(row) > 3  else "",
                    "Total Received":    row[4]  if len(row) > 4  else "",
                    "Current Inventory": row[5]  if len(row) > 5  else "",
                    "Storage Box Qty":   row[6]  if len(row) > 6  else "",
                    "Total Output":      row[7]  if len(row) > 7  else "",
                    "Avg Units/Day":     row[8]  if len(row) > 8  else "",
                    "Days Available":    row[9]  if len(row) > 9  else "",
                    "Refill Qty (pcs)":  row[10] if len(row) > 10 else "",
                    "Refill Qty (ctn)":  row[11] if len(row) > 11 else "",
                })

            df = pd.DataFrame(inv_data)
            st.dataframe(df, use_container_width=True, hide_index=True, height=(len(inv_data) + 1) * 35 + 3)

            if role == "admin":
                st.divider()
                try:
                    rcv_excel = sheet_to_excel("Inventory Received")
                    st.download_button(
                        label     = "⬇️ Download Inventory Received",
                        data      = rcv_excel,
                        file_name = "inventory_received.xlsx",
                        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Could not prepare Inventory Received download: {e}")

                st.divider()
                st.subheader("🔄 Push Inventory to Faire")
                st.caption("⚠️ TEST MODE: Pushes only T008-WM-MNY with quantity 99 to verify API connection.")

                if "push_inv_confirm" not in st.session_state:
                    st.session_state["push_inv_confirm"] = False

                if not st.session_state["push_inv_confirm"]:
                    if st.button("🔄 Push Inventory to Faire", type="primary"):
                        st.session_state["push_inv_confirm"] = True
                        st.rerun()
                else:
                    st.warning("⚠️ Are you sure? This will update inventory levels in Faire. This cannot be undone.")
                    push_col1, push_col2 = st.columns([1, 1])
                    with push_col1:
                        if st.button("✅ Yes, Push Inventory", type="primary"):
                            with st.spinner("Pushing inventory to Faire..."):
                                try:
                                    test_payload = [{"sku": "TUG-WM-MNY", "on_hand_quantity": 99}]
                                    result = push_inventory_to_faire(test_payload)
                                    st.session_state["push_inv_confirm"] = False
                                    st.success("✅ Test successful! TUG-WM-MNY set to 99 in Faire. Check your Faire account to verify.")
                                    st.json(result)
                                except Exception as e:
                                    st.error(f"Failed to push inventory: {e}")
                                    st.session_state["push_inv_confirm"] = False
                    with push_col2:
                        if st.button("❌ Cancel", key="cancel_push"):
                            st.session_state["push_inv_confirm"] = False
                            st.rerun()

    except Exception as e:
        st.error(f"Could not load inventory: {e}")
        st.info("Tip: Make sure the tab is named exactly 'Inventory' and the service account has Editor access to the sheet.")


# ─────────────────────────────────────────────
# PAGE: WSP ORDERS (Admin only)
# ─────────────────────────────────────────────
elif page == "🛒 WSP Orders":
    st.header("🛒 WholesalePet.com Orders")
    st.caption("Admin only. Enter new WSP orders below.")



    # View existing WSP orders
    try:

        wsp_orders_view = get_wsp_orders()
        if wsp_orders_view:
            st.subheader("Existing WSP Orders")

            orders_dict = {}
            for o in wsp_orders_view:
                order_num = o["order_number"]
                orders_dict[order_num] = {
                    "Order Date":    o["created_at"],
                    "Customer":      o["customer"],
                    "PDF":           "✅" if o.get("drive_file_id") else "—",
                    "drive_file_id": o.get("drive_file_id", ""),
                }
                sku_lookup = {item["sku"]: item["quantity"] for item in o["items"]}
                for sku in ALL_SKUS:
                    orders_dict[order_num][sku] = sku_lookup.get(sku, "")

            order_nums = list(orders_dict.keys())
            table_rows = []

            for meta in ["Order Date", "Customer", "PDF"]:
                row_data = {"SKU": meta}
                for o in order_nums:
                    row_data[o] = orders_dict[o].get(meta, "")
                table_rows.append(row_data)

            table_rows.append({"SKU": "─" * 10})

            for sku in ALL_SKUS:
                row_data = {"SKU": sku}
                for o in order_nums:
                    row_data[o] = orders_dict[o].get(sku, "")
                table_rows.append(row_data)

            df = pd.DataFrame(table_rows)
            st.dataframe(df, use_container_width=True, hide_index=True,
                         height=(len(table_rows) + 1) * 35 + 3)

            # Delete an order
            st.divider()
            st.subheader("🗑️ Delete a WSP Order")
            order_to_delete = st.selectbox("Select order to delete", options=["— select —"] + order_nums)
            if order_to_delete != "— select —":
                st.warning(f"You are about to delete order **{order_to_delete}** ({orders_dict[order_to_delete]['Customer']} — {orders_dict[order_to_delete]['Order Date']}). This cannot be undone.")
                if st.button("🗑️ Confirm Delete", type="primary"):
                    try:
                        ws_del   = get_sheet("WSP Orders")
                        data_del = ws_del.get_all_values()
                        col_index = None
                        # Order # is in row index 1 (second row), find matching column
                        if len(data_del) > 1:
                            for col_i, val in enumerate(data_del[1]):
                                if val == order_to_delete:
                                    col_index = col_i + 1  # gspread is 1-indexed
                                    break
                        if col_index:
                            file_id = orders_dict[order_to_delete].get("drive_file_id", "")
                            if file_id:
                                delete_pdf_from_sheet(file_id)
                            ws_del.delete_columns(col_index)
                            st.success(f"✅ Order {order_to_delete} deleted successfully!")
                            st.rerun()
                        else:
                            st.error("Could not find that order in the sheet.")
                    except Exception as e:
                        st.error(f"Failed to delete order: {e}")
        else:
            st.info("No WSP orders entered yet.")

    except Exception as e:
        st.error(f"Could not load WSP orders: {e}")
        wsp_orders_view = []

    st.divider()
    st.subheader("➕ Enter New WSP Order")

    with st.form("wsp_order_form"):
        col1, col2, col3 = st.columns(3)
        with col1:
            order_date = st.date_input("Order Date")
        with col2:
            order_num = st.text_input("Order #")
        with col3:
            customer = st.text_input("Customer", value="")

        uploaded_pdf = st.file_uploader("Upload Packing Slip PDF (optional)", type=["pdf"])

        st.markdown("**Enter quantities for each SKU (leave at 0 if not ordered):**")
        quantities = {}
        for sku in ALL_SKUS:
            quantities[sku] = st.number_input(sku, min_value=0, value=0, step=1, key=f"wsp_{sku}")

        submitted = st.form_submit_button("💾 Save WSP Order")

        if submitted:
            if not order_num:
                st.error("Please enter an Order #.")
            else:
                try:
                    # Upload PDF to Google Drive if provided
                    drive_file_id = ""
                    if uploaded_pdf is not None:
                        if not DRIVE_FOLDER_ID:
                            st.session_state["wsp_debug"] = "ERROR: DRIVE_FOLDER_ID is not set in Streamlit secrets."
                        else:
                            try:
                                customer_safe = customer.replace("/", "-").replace("\\", "-")
                                pdf_filename  = f"{order_num}_{customer_safe}_PackingSlip.pdf"
                                pdf_data      = uploaded_pdf.getvalue()
                                drive_file_id = store_pdf_in_sheet(pdf_data, pdf_filename)
                            except Exception as pdf_err:
                                st.session_state["wsp_debug"] = f"❌ PDF upload failed: {pdf_err}"

                    # Save order to Google Sheets as a new column
                    # Layout: Row1=Date, Row2=Order#, Row3=Customer, Row4=DriveFileID, Row5+=SKUs
                    ws   = get_sheet("WSP Orders")
                    data = ws.get_all_values()
                    # Find next empty column by checking max cols across all rows
                    max_cols = max((len(row) for row in data), default=0)
                    next_col = max_cols + 1

                    # Build the column values in order
                    col_values = (
                        [str(order_date), order_num, customer, drive_file_id] +
                        [quantities[sku] if quantities[sku] > 0 else "" for sku in ALL_SKUS]
                    )

                    # Write each value to the correct row in the next column
                    from gspread.utils import rowcol_to_a1
                    cell_updates = []
                    for row_idx, val in enumerate(col_values, start=1):
                        cell_updates.append({
                            "range": rowcol_to_a1(row_idx, next_col),
                            "values": [[val]],
                        })
                    ws.batch_update(cell_updates)
                    st.success(f"✅ Order {order_num} saved!" + (" PDF uploaded." if drive_file_id else ""))
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to save order: {e}")


# ─────────────────────────────────────────────
# PAGE: SHIPPING INFO
# ─────────────────────────────────────────────
elif page == "📦 Shipping Info":
    st.header("📦 Shipping Info")
    st.caption("Enter carton dimensions and SKU quantities for each order.")

    # Get all current orders
    faire_orders  = st.session_state.get("faire_orders", [])
    wsp_orders    = get_wsp_orders()
    all_orders    = faire_orders + wsp_orders

    if not all_orders:
        st.info("No orders found. Go to the Orders page and click '🔄 Refresh from Faire' first.")
        st.stop()

    # Order selector
    order_options = {
        f"{o['order_number']} — {o['customer']}": o for o in all_orders
    }
    selected_label = st.selectbox("Select Order", options=list(order_options.keys()))
    selected_order = order_options[selected_label]
    order_num      = selected_order["order_number"]

    # Build ordered SKU list for this order
    ordered_skus = {item["sku"]: item["quantity"] for item in selected_order["items"]}

    st.divider()

    # Load existing cartons for this order
    existing_cartons = get_shipping_info(order_num)
    assigned_qtys    = {sku: 0 for sku in ordered_skus}
    for c in existing_cartons.values():
        for sku, qty in c["skus"].items():
            if sku in assigned_qtys:
                assigned_qtys[sku] += qty

    # Show remaining quantities
    st.subheader(f"📋 {order_num} — SKU Summary")
    summary_data = []
    all_assigned = True
    for sku, total_qty in ordered_skus.items():
        assigned  = assigned_qtys.get(sku, 0)
        remaining = total_qty - assigned
        if remaining != 0:
            all_assigned = False
        summary_data.append({
            "SKU":       sku,
            "Ordered":   total_qty,
            "Assigned":  assigned,
            "Remaining": remaining,
        })

    import pandas as pd
    summary_df = pd.DataFrame(summary_data)
    st.dataframe(summary_df, use_container_width=True, hide_index=True,
                 height=(len(summary_data) + 1) * 35 + 3)

    if all_assigned:
        st.success("✅ All SKUs have been assigned to cartons!")
    else:
        remaining_count = sum(1 for r in summary_data if r["Remaining"] != 0)
        st.warning(f"⚠️ {remaining_count} SKU(s) still have unassigned quantities.")

    st.divider()

    # Show existing cartons
    if existing_cartons:
        st.subheader("🗃️ Existing Cartons")
        for c_num in sorted(existing_cartons.keys()):
            c = existing_cartons[c_num]
            total_weight = sum(
                get_sku_weight(sku) * qty
                for sku, qty in c["skus"].items()
            )
            with st.expander(
                f"Carton {c_num} — {c['length']}\" × {c['width']}\" × {c['height']}\"  |  "
                f"Weight: {round(total_weight, 2)} lbs",
                expanded=False
            ):
                for sku, qty in c["skus"].items():
                    w = round(get_sku_weight(sku) * qty, 3)
                    st.write(f"**{sku}**: {qty} units ({w} lbs)")

                if st.button(f"🗑️ Delete Carton {c_num}", key=f"del_carton_{c_num}"):
                    delete_carton(order_num, c_num)
                    st.success(f"Carton {c_num} deleted.")
                    st.rerun()

    st.divider()

    # Only show Add Carton form if there are still unassigned SKUs
    if all_assigned:
        st.stop()

    # Add new carton
    next_carton_num = max(existing_cartons.keys(), default=0) + 1
    st.subheader(f"➕ Add Carton {next_carton_num}")

    with st.form(f"carton_form_{order_num}_{next_carton_num}"):
        st.markdown("**Dimensions (inches):**")
        dim_col1, dim_col2, dim_col3 = st.columns(3)
        with dim_col1:
            length = st.number_input('Length "', min_value=0.0, step=0.1)
        with dim_col2:
            width  = st.number_input('Width "',  min_value=0.0, step=0.1)
        with dim_col3:
            height = st.number_input('Height "', min_value=0.0, step=0.1)

        st.markdown("**SKU Quantities for this carton:**")
        st.caption("Only SKUs with remaining quantities are shown.")

        sku_qtys = {}
        available_skus = {
            sku: qty for sku, qty in ordered_skus.items()
            if (qty - assigned_qtys.get(sku, 0)) > 0
        }

        if not available_skus:
            st.success("All SKUs are fully assigned!")
        else:
            for sku, total_qty in available_skus.items():
                remaining = total_qty - assigned_qtys.get(sku, 0)
                label     = f"{sku}  (remaining: {remaining})"
                sku_qtys[sku] = st.number_input(
                    label,
                    min_value = 0,
                    max_value = int(remaining),
                    value     = 0,
                    step      = 1,
                    key       = f"carton_{next_carton_num}_{sku}",
                )

        # Live weight preview
        preview_weight = sum(
            get_sku_weight(sku) * qty for sku, qty in sku_qtys.items()
        )
        st.info(f"📦 Estimated carton weight: **{round(preview_weight, 2)} lbs**")

        submitted = st.form_submit_button("💾 Save Carton")
        if submitted:
            if length == 0 or width == 0 or height == 0:
                st.error("Please enter all dimensions.")
            elif sum(sku_qtys.values()) == 0:
                st.error("Please assign at least one SKU to this carton.")
            else:
                try:
                    save_carton(order_num, next_carton_num, length, width, height, sku_qtys)
                    st.success(f"✅ Carton {next_carton_num} saved! Weight: {round(preview_weight, 2)} lbs")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to save carton: {e}")


# ─────────────────────────────────────────────
# INVENTORY PUSH TEST (Admin only)
# Added at bottom of Inventory page via session flag
# ─────────────────────────────────────────────
